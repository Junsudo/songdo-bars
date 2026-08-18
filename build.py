# -*- coding: utf-8 -*-
# 송도 술집 후보 표 빌더
# 입력: data/yeonsu_restaurants.csv (인허가, UTF-8 변환본), data/diningcode_all.json,
#       data/triplestreet.json, data/canalwalk_liquor.json, data/dailybeer.html
# 출력: 송도_술집_후보.xlsx
import csv, json, re, unicodedata
from collections import defaultdict

DATA = "data"
PUB_TYPES = {"호프/통닭", "감성주점", "정종/대포집/소주방"}

def norm(s):
    if not s: return ""
    s = unicodedata.normalize("NFKC", str(s)).lower()
    s = re.sub(r"\(주\)|주식회사|㈜", "", s)
    return re.sub(r"[^0-9a-z가-힣]", "", s)

def base_name(s):
    # 지점 접미사 제거: "...송도점", "...인천송도점", "...트리플스트리트점" 등
    n = norm(s)
    n = re.sub(r"(인천)?(송도)?(국제도시|타임스페이스|트리플스트리트|커낼워크|센트럴파크|워터프런트|스마트스퀘어|롯데몰|캠퍼스타운역?)?(\d공구)?점$", "", n)
    return n or norm(s)

def lot_of(addr):
    m = re.search(r"송도동\s*(\d+(?:-\d+)?)", addr or "")
    return m.group(1) if m else None

# ── 1. 인허가 ──────────────────────────────────────────────
rows = list(csv.DictReader(open(f"{DATA}/yeonsu_restaurants.csv", encoding="utf-8")))
alive = [r for r in rows if r["영업상태명"] == "영업/정상"]
songdo = [r for r in alive if "송도동" in (r["지번주소"] or "")]
pubs = [r for r in songdo if r["업태구분명"] in PUB_TYPES]
print(f"인허가: 전체 {len(rows)}, 영업중 {len(alive)}, 송도 {len(songdo)}, 주점류 {len(pubs)}")

def area_of(r):
    try: return float(r["소재지면적"])
    except (ValueError, TypeError): return 0.0

# 송도 전체 영업중 lookup (주점류 외 업태로 등록된 술집 회수용)
by_name = defaultdict(list); by_base = defaultdict(list); by_lot = defaultdict(list)
for r in songdo:
    by_name[norm(r["사업장명"])].append(r)
    by_base[base_name(r["사업장명"])].append(r)
    lot = lot_of(r["지번주소"])
    if lot: by_lot[lot].append(r)

# ── 2. 다이닝코드 매칭 ─────────────────────────────────────
dc = json.load(open(f"{DATA}/diningcode_all.json"))
print(f"다이닝코드: {len(dc)}")

def match_dc(poi):
    cands = []
    full = norm(poi["nm"] + (poi.get("branch") or ""))
    nm = norm(poi["nm"]); bn = base_name(poi["nm"])
    for key, table, how in ((full, by_name, "이름+지점"), (nm, by_name, "이름"), (bn, by_base, "기본이름")):
        if key and table.get(key):
            cands = table[key]; break
    else:
        how = None
        lot = lot_of(poi.get("addr") or poi.get("road_addr") or "")
        if lot:
            for r in by_lot.get(lot, []):
                rb, pb = base_name(r["사업장명"]), bn
                if pb and rb and len(pb) >= 2 and (pb in rb or rb in pb):
                    cands = [r]; how = "지번+이름유사"; break
    if not cands: return None, None
    return max(cands, key=area_of), how

dc_match = {}   # v_rid -> (인허가 row, method)
lic_dc = {}     # 인허가 관리번호 -> poi
for poi in dc:
    r, how = match_dc(poi)
    if r:
        dc_match[poi["v_rid"]] = (r, how)
        lic_dc.setdefault(r["관리번호"], poi)
print(f"DC→인허가 매칭: {len(dc_match)}/{len(dc)}")

# ── 2b. 카카오 Local API 술집 매칭 ─────────────────────────
kakao = json.load(open(f"{DATA}/kakao_bars.json"))
kakao_match = {}   # kakao id -> (인허가 row, method)
lic_kakao = {}     # 인허가 관리번호 -> kakao place
for kp in kakao:
    pseudo = {"nm": kp["place_name"], "branch": "",
              "addr": kp.get("address_name"), "road_addr": kp.get("road_address_name")}
    r, how = match_dc(pseudo)
    if r:
        kakao_match[kp["id"]] = (r, how)
        lic_kakao.setdefault(r["관리번호"], kp)
print(f"카카오→인허가 매칭: {len(kakao_match)}/{len(kakao)}")

# ── 3. 디렉토리 ────────────────────────────────────────────
ts = json.load(open(f"{DATA}/triplestreet.json"))
ts_late = [x for x in ts if str(x.get("closeTime", "")).startswith("익일")]
cw = json.load(open(f"{DATA}/canalwalk_liquor.json"))
db_html = open(f"{DATA}/dailybeer.html", encoding="utf-8", errors="replace").read()
m = re.search(r"var\s+slMapData\s*=\s*(\{.*?\});", db_html, re.S)
db_stores = []
if m:
    try:
        blob = json.loads(m.group(1))
        def flat(v):
            if isinstance(v, dict):
                for vv in v.values(): yield from flat(vv)
            elif isinstance(v, list):
                for vv in v: yield from flat(vv)
        for item in flat(blob):
            pass
    except Exception:
        blob = None
    # slMapData 구조가 중첩이면 name/address 쌍을 정규식으로 직접 회수
raw = re.findall(r'\{[^{}]*"name"[^{}]*\}', db_html.replace("\\/", "/"))
for r_ in raw:
    try:
        o = json.loads(r_)
        if "송도" in json.dumps(o, ensure_ascii=False):
            db_stores.append(o)
    except Exception: pass
print(f"트리플 심야 {len(ts_late)}, 커낼워크 주류 {len(cw)}, 생활맥주 송도 {len(db_stores)}")

dir_names = {base_name(x["name"]) for x in ts_late} | {base_name(x["name"]) for x in cw} | {base_name(x.get("name","")) for x in db_stores}

def dir_flag(r):
    b = base_name(r["사업장명"])
    return "Y" if b in dir_names else ""

# ── 4. 워크북 ──────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
ARIAL = "Arial"
HDR = Font(name=ARIAL, bold=True, size=10)
CELL = Font(name=ARIAL, size=10)
YELLOW = PatternFill("solid", fgColor="FFF2CC")

def style_sheet(ws, widths, n_rows):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.font = HDR
    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
        for c in row:
            if c.font is None or c.font.name != ARIAL:
                c.font = CELL
    ws.freeze_panes = "A2"

def clean(s):
    return re.sub(r"<[^>]+>", "", str(s or "")).strip()

def kw_text(poi):
    kw = poi.get("keyword") or []
    if isinstance(kw, str): return clean(kw)
    out = []
    for k in kw[:4]:
        if isinstance(k, str): out.append(clean(k))
        elif isinstance(k, dict):
            v = k.get("keyword") or k.get("nm") or next((x for x in k.values() if isinstance(x, str)), "")
            if v: out.append(clean(v))
    return ", ".join(x for x in out if x)

def kakao_flag(r):
    kp = lic_kakao.get(r["관리번호"])
    if not kp: return ""
    parts = (kp.get("category_name") or "").split(" > ")
    return parts[-1] if parts and parts[-1] else "Y"

def dc_cols(r):
    poi = lic_dc.get(r["관리번호"])
    if not poi: return ["", "", "", "", ""]
    return [clean(poi.get("category")), poi.get("user_score") or "", poi.get("review_cnt") or "",
            clean(poi.get("open_status")), kw_text(poi)]

HEADERS = ["사업장명", "업태구분명", "소재지면적(㎡)", "추정좌석(보수 2.5㎡/석)", "추정좌석(낙관 2.0㎡/석)",
           "전화번호", "도로명주소", "지번주소", "인허가일자", "다중이용업소", "디렉토리확인", "카카오확인",
           "DC카테고리", "DC평점", "DC리뷰수", "DC영업표시(수집시각)", "DC키워드"]
WIDTHS = [24, 14, 13, 12, 12, 14, 42, 34, 11, 10, 10, 14, 16, 8, 8, 10, 26]

def write_rows(ws, rs):
    ws.append(HEADERS)
    for i, r in enumerate(rs, start=2):
        ws.append([r["사업장명"], r["업태구분명"], area_of(r) or None,
                   f"=IF(C{i}>0,ROUND(C{i}/2.5,0),\"\")", f"=IF(C{i}>0,ROUND(C{i}/2,0),\"\")",
                   r["전화번호"] or "", r["도로명주소"] or "", r["지번주소"] or "",
                   r["인허가일자"] or "", r["다중이용업소여부"] or "", dir_flag(r), kakao_flag(r)] + dc_cols(r))
        ws.cell(row=i, column=3).number_format = "0.0"
        if area_of(r) >= 250:
            for c in ws[i][:5]: c.fill = YELLOW
    style_sheet(ws, WIDTHS, len(rs))

# 시트1: 요약
s = wb.active; s.title = "요약"
summary = [
    ["송도(인천 연수구 송도동) 술집 후보 표", ""],
    ["생성일", "2026-08-18"],
    ["", ""],
    ["데이터 소스", "내용"],
    ["행정안전부 인허가 데이터(연수구, 2026-08-18 다운로드)", "존재·영업상태·소재지면적·업태·전화의 1차 근거. file.localdata.go.kr orgCode=3520000"],
    ["다이닝코드 '송도 술집' 371건(2026-08-18 수집)", "카테고리·평점·리뷰수·영업표시. 정렬 4종+거리순 앵커 10곳 유니온으로 전수 수집"],
    ["카카오 Local API FD6 quadtree(2026-08-18, 406호출)", "송도동 술집 카테고리 171곳 — 카카오확인 컬럼과 카카오맵(술집) 시트"],
    ["트리플스트리트 공식 매장 API", "심야(익일 마감) 매장 → 디렉토리확인 플래그"],
    ["커낼워크 상인회(주류 업종 5곳)·생활맥주 본사 매장 데이터", "디렉토리확인 플래그"],
    ["", ""],
    ["가정", "좌석수는 어떤 공식 데이터에도 없음. 추정좌석 = 소재지면적 ÷ 2.5㎡(보수) 또는 2.0㎡(낙관). 소재지면적은 인허가 신고 면적(주방·통로 포함)이라 실제 홀 좌석과 차이 가능"],
    ["한계(정확성)", "소재지면적은 인허가 시점 신고값 — 증축·축소 미반영 가능. 다이닝코드 영업표시·평점은 리뷰 수가 적어(대부분 10건 미만) 참고용. 최종 좌석·단체·대관은 전화 확인 필요"],
    ["제외", "유흥주점·단란주점(행사 성격 불일치, 송도동 단란주점 0곳), 대관 실적 역추적(요청으로 제외)"],
    ["", ""],
    ["시트 안내", ""],
    ["주점류(면적순)", "인허가 업태가 주점류(호프/통닭·감성주점·정종/대포집/소주방)인 송도동 영업중 업소 전체, 면적 내림차순. 노란색 = 250㎡ 이상(≈100석급)"],
    ["100석후보", "면적 200㎡ 이상 + 술집 성격(주점류 업태 또는 다이닝코드 '송도 술집' 결과에 매칭)"],
    ["다이닝코드371", "다이닝코드 검색 결과 전체와 인허가 매칭(면적·업태 회수)"],
    ["카카오맵(술집)", "카카오 수집 171곳 전체와 인허가 매칭 — 인허가에 없는 카카오 단독 항목은 상호 변경·신규 개업·타 업태 등록 가능성"],
    ["디렉토리", "트리플스트리트 심야 매장·커낼워크 주류·생활맥주 원본"],
]
for row in summary: s.append(row)
s.column_dimensions["A"].width = 44; s.column_dimensions["B"].width = 100
for row in s.iter_rows():
    for c in row: c.font = CELL

for r_ in (s[1], s[4], s[14]):
    for c in r_: c.font = HDR
for row in s.iter_rows():
    for c in row: c.alignment = Alignment(wrap_text=True, vertical="top")

# 시트2: 주점류(면적순)
ws = wb.create_sheet("주점류(면적순)")
pubs_sorted = sorted(pubs, key=area_of, reverse=True)
write_rows(ws, pubs_sorted)

# 시트3: 100석후보
ws = wb.create_sheet("100석후보")
barlike = [r for r in songdo if (r["업태구분명"] in PUB_TYPES or r["관리번호"] in lic_dc or r["관리번호"] in lic_kakao) and area_of(r) >= 200]
write_rows(ws, sorted(barlike, key=area_of, reverse=True))
print(f"100석후보(200㎡+ bar-like): {len(barlike)}")

# 시트4: 다이닝코드371
ws = wb.create_sheet("다이닝코드371")
ws.append(["상호(다이닝코드)", "지점", "카테고리", "평점", "리뷰수", "영업표시(수집시각)", "전화", "도로명주소(DC)",
           "인허가매칭", "매칭방법", "인허가 사업장명", "업태구분명", "소재지면적(㎡)", "인허가 전화"])
dc_sorted = sorted(dc, key=lambda p: (-(float(p.get("user_score") or 0)), -(p.get("review_cnt") or 0)))
for i, poi in enumerate(dc_sorted, start=2):
    mr = dc_match.get(poi["v_rid"])
    row = [poi.get("nm"), poi.get("branch") or "", clean(poi.get("category")), poi.get("user_score") or "",
           poi.get("review_cnt") or "", clean(poi.get("open_status")), poi.get("phone") or "",
           poi.get("road_addr") or poi.get("addr") or ""]
    if mr:
        r, how = mr
        row += ["Y", how, r["사업장명"], r["업태구분명"], area_of(r) or None, r["전화번호"] or ""]
    else:
        row += ["", "", "", "", None, ""]
    ws.append(row)
    ws.cell(row=i, column=13).number_format = "0.0"
style_sheet(ws, [22, 12, 14, 7, 8, 10, 14, 40, 9, 12, 22, 14, 13, 14], len(dc_sorted))

# 시트5: 카카오맵(술집)
ws = wb.create_sheet("카카오맵(술집)")
ws.append(["상호(카카오)", "세부카테고리", "전화", "지번주소(카카오)", "place_url", "인허가매칭", "매칭방법",
           "인허가 사업장명", "업태구분명", "소재지면적(㎡)", "DC에도있음"])
dc_name_set = {norm(p["nm"]) for p in dc}
def kmatch_area(kp):
    mr = kakao_match.get(kp["id"])
    return area_of(mr[0]) if mr else -1.0
for i, kp in enumerate(sorted(kakao, key=kmatch_area, reverse=True), start=2):
    mr = kakao_match.get(kp["id"])
    parts = (kp.get("category_name") or "").split(" > ")
    row = [kp["place_name"], parts[-1] if parts else "", kp.get("phone") or "",
           kp.get("address_name") or "", kp.get("place_url") or ""]
    if mr:
        r, how = mr
        dcflag = "Y" if r["관리번호"] in lic_dc or norm(kp["place_name"]) in dc_name_set else ""
        row += ["Y", how, r["사업장명"], r["업태구분명"], area_of(r) or None, dcflag]
    else:
        dcflag = "Y" if norm(kp["place_name"]) in dc_name_set else ""
        row += ["", "", "", "", None, dcflag]
    ws.append(row)
    ws.cell(row=i, column=10).number_format = "0.0"
style_sheet(ws, [24, 13, 14, 30, 34, 9, 12, 24, 14, 13, 9], len(kakao))

# 시트6: 디렉토리
ws = wb.create_sheet("디렉토리")
ws.append(["출처", "매장명", "위치", "전화", "영업시간", "태그/업종"])
for x in ts_late:
    ws.append(["트리플스트리트(심야)", x["name"], x.get("location") or "", x.get("tel") or "",
               f"{x.get('openTime','')}~{x.get('closeTime','')}", ", ".join(x.get("tags") or [])])
for x in cw:
    ws.append(["커낼워크", x["name"], f"{x['building']} {x['dong']} {x['ho']}", "", "", x["type"]])
for x in db_stores:
    ws.append(["생활맥주", x.get("name",""), x.get("address",""), x.get("tel",""), "", "프랜차이즈"])
style_sheet(ws, [18, 26, 30, 14, 22, 30], len(ts_late) + len(cw) + len(db_stores))

wb.save("송도_술집_후보.xlsx")
print("저장 완료: 송도_술집_후보.xlsx")
